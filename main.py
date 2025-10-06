"""
Repair Tracker - FastAPI application
"""
from datetime import datetime, timedelta
from typing import List, Optional
from fastapi import FastAPI, HTTPException, Depends, status, Request, Form, File, UploadFile
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.security import HTTPBearer
from fastapi.middleware.cors import CORSMiddleware
import sqlite3
import logging
import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import io
import chardet
from fastapi.responses import StreamingResponse

from database import get_db_connection, init_database
from models import (
    UserCreate, UserLogin, UserResponse, AssetCreate, AssetUpdate, AssetResponse,
    RepairCreate, RepairUpdate, RepairResponse, RepairFilter, AssetFilter,
    Token, UserRole, RepairStatus, UserProfileUpdate, PasswordChange, 
    UserSettings, UserSettingsResponse, ImportResult, AssetImportRow, RepairImportRow
)
from auth import (
    authenticate_user, create_access_token, get_current_user, get_current_admin_user,
    get_password_hash, verify_password, ACCESS_TOKEN_EXPIRE_MINUTES
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize FastAPI app
app = FastAPI(title="Repair Tracker", version="1.0.0")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify your frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# Templates
templates = Jinja2Templates(directory="templates")

# Helper function to detect file encoding
def detect_encoding(content):
    """Detect file encoding and decode content"""
    encodings = ['utf-8', 'utf-8-sig', 'cp1251', 'windows-1251', 'latin1']
    
    # Try to detect encoding using chardet
    detected = chardet.detect(content)
    if detected['encoding'] and detected['confidence'] > 0.7:
        encodings.insert(0, detected['encoding'])
    
    for encoding in encodings:
        try:
            return content.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    
    raise ValueError("Unable to decode file with any supported encoding")

# Initialize database on startup
@app.on_event("startup")
async def startup_event():
    init_database()

# Authentication endpoints
@app.post("/auth/register", response_model=UserResponse)
async def register(user: UserCreate):
    """Register a new user"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if user already exists
        cursor.execute("SELECT id FROM users WHERE email = ?", (user.email,))
        if cursor.fetchone():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Email already registered"
            )
        
        # Create new user
        password_hash = get_password_hash(user.password)
        cursor.execute(
            "INSERT INTO users (email, name, password_hash) VALUES (?, ?, ?)",
            (user.email, user.name, password_hash)
        )
        conn.commit()
        
        # Get created user
        user_id = cursor.lastrowid
        cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
        user_row = cursor.fetchone()
        
        return UserResponse(
            id=user_row["id"],
            email=user_row["email"],
            name=user_row["name"],
            role=user_row["role"],
            created_at=datetime.fromisoformat(user_row["created_at"]),
            updated_at=datetime.fromisoformat(user_row["updated_at"])
        )
    except Exception as e:
        conn.rollback()
        logger.error(f"Registration error: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Registration failed")
    finally:
        conn.close()

@app.post("/auth/login", response_model=Token)
async def login(user_credentials: UserLogin):
    """Login user and return JWT token"""
    user = authenticate_user(user_credentials.email, user_credentials.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.email}, expires_delta=access_token_expires
    )
    
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/auth/me", response_model=UserResponse)
async def get_current_user_info(current_user: UserResponse = Depends(get_current_user)):
    """Get current user information"""
    return current_user

# Asset endpoints
@app.get("/api/assets", response_model=List[AssetResponse])
async def get_assets(
    current_user: UserResponse = Depends(get_current_user),
    owner_id: Optional[int] = None,
    type: Optional[str] = None,
    search: Optional[str] = None
):
    """Get assets with optional filtering"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        query = "SELECT * FROM assets WHERE 1=1"
        params = []
        
        # Apply role-based filtering
        if current_user.role != "ADMIN":
            query += " AND owner_id = ?"
            params.append(current_user.id)
        elif owner_id is not None:
            query += " AND owner_id = ?"
            params.append(owner_id)
        
        # Apply additional filters
        if type:
            query += " AND type = ?"
            params.append(type)
        
        if search:
            query += " AND name LIKE ?"
            params.append(f"%{search}%")
        
        query += " ORDER BY created_at DESC"
        
        cursor.execute(query, params)
        assets = []
        
        for row in cursor.fetchall():
            assets.append(AssetResponse(
                id=row["id"],
                name=row["name"],
                type=row["type"],
                owner_id=row["owner_id"],
                created_at=datetime.fromisoformat(row["created_at"]),
                updated_at=datetime.fromisoformat(row["updated_at"])
            ))
        
        return assets
    finally:
        conn.close()

@app.post("/api/assets", response_model=AssetResponse)
async def create_asset(
    asset: AssetCreate,
    current_user: UserResponse = Depends(get_current_user)
):
    """Create a new asset"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute(
            "INSERT INTO assets (name, type, owner_id) VALUES (?, ?, ?)",
            (asset.name, asset.type, current_user.id)
        )
        conn.commit()
        
        asset_id = cursor.lastrowid
        cursor.execute("SELECT * FROM assets WHERE id = ?", (asset_id,))
        asset_row = cursor.fetchone()
        
        return AssetResponse(
            id=asset_row["id"],
            name=asset_row["name"],
            type=asset_row["type"],
            owner_id=asset_row["owner_id"],
            created_at=datetime.fromisoformat(asset_row["created_at"]),
            updated_at=datetime.fromisoformat(asset_row["updated_at"])
        )
    except Exception as e:
        conn.rollback()
        logger.error(f"Asset creation error: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Asset creation failed")
    finally:
        conn.close()

@app.get("/api/assets/{asset_id}", response_model=AssetResponse)
async def get_asset(
    asset_id: int,
    current_user: UserResponse = Depends(get_current_user)
):
    """Get asset by ID"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        query = "SELECT * FROM assets WHERE id = ?"
        params = [asset_id]
        
        if current_user.role != "ADMIN":
            query += " AND owner_id = ?"
            params.append(current_user.id)
        
        cursor.execute(query, params)
        asset_row = cursor.fetchone()
        
        if not asset_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")
        
        return AssetResponse(
            id=asset_row["id"],
            name=asset_row["name"],
            type=asset_row["type"],
            owner_id=asset_row["owner_id"],
            created_at=datetime.fromisoformat(asset_row["created_at"]),
            updated_at=datetime.fromisoformat(asset_row["updated_at"])
        )
    finally:
        conn.close()

@app.put("/api/assets/{asset_id}", response_model=AssetResponse)
async def update_asset(
    asset_id: int,
    asset_update: AssetUpdate,
    current_user: UserResponse = Depends(get_current_user)
):
    """Update asset"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if asset exists and user has permission
        query = "SELECT * FROM assets WHERE id = ?"
        params = [asset_id]
        
        if current_user.role != "ADMIN":
            query += " AND owner_id = ?"
            params.append(current_user.id)
        
        cursor.execute(query, params)
        asset_row = cursor.fetchone()
        
        if not asset_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")
        
        # Update asset
        update_fields = []
        update_params = []
        
        if asset_update.name is not None:
            update_fields.append("name = ?")
            update_params.append(asset_update.name)
        
        if asset_update.type is not None:
            update_fields.append("type = ?")
            update_params.append(asset_update.type)
        
        if update_fields:
            update_fields.append("updated_at = ?")
            update_params.append(datetime.utcnow().isoformat())
            update_params.append(asset_id)
            
            cursor.execute(
                f"UPDATE assets SET {', '.join(update_fields)} WHERE id = ?",
                update_params
            )
            conn.commit()
        
        # Return updated asset
        cursor.execute("SELECT * FROM assets WHERE id = ?", (asset_id,))
        updated_asset = cursor.fetchone()
        
        return AssetResponse(
            id=updated_asset["id"],
            name=updated_asset["name"],
            type=updated_asset["type"],
            owner_id=updated_asset["owner_id"],
            created_at=datetime.fromisoformat(updated_asset["created_at"]),
            updated_at=datetime.fromisoformat(updated_asset["updated_at"])
        )
    except Exception as e:
        conn.rollback()
        logger.error(f"Asset update error: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Asset update failed")
    finally:
        conn.close()

@app.delete("/api/assets/{asset_id}")
async def delete_asset(
    asset_id: int,
    current_user: UserResponse = Depends(get_current_user)
):
    """Delete asset"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if asset exists and user has permission
        query = "SELECT * FROM assets WHERE id = ?"
        params = [asset_id]
        
        if current_user.role != "ADMIN":
            query += " AND owner_id = ?"
            params.append(current_user.id)
        
        cursor.execute(query, params)
        asset_row = cursor.fetchone()
        
        if not asset_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")
        
        # Delete asset (repairs will be deleted by CASCADE)
        cursor.execute("DELETE FROM assets WHERE id = ?", (asset_id,))
        conn.commit()
        
        return {"message": "Asset deleted successfully"}
    except Exception as e:
        conn.rollback()
        logger.error(f"Asset deletion error: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Asset deletion failed")
    finally:
        conn.close()

# Repair endpoints
@app.get("/api/repairs", response_model=List[RepairResponse])
async def get_repairs(
    current_user: UserResponse = Depends(get_current_user),
    property_id: Optional[int] = None,
    status: Optional[RepairStatus] = None,
    year: Optional[int] = None,
    sort_by: Optional[str] = "date",
    sort_order: Optional[str] = "desc"
):
    """Get repairs with filtering and sorting"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        query = """
            SELECT r.* FROM repairs r
            JOIN assets a ON r.property_id = a.id
            WHERE 1=1
        """
        params = []
        
        # Apply role-based filtering
        if current_user.role != "ADMIN":
            query += " AND a.owner_id = ?"
            params.append(current_user.id)
        
        # Apply filters
        if property_id:
            query += " AND r.property_id = ?"
            params.append(property_id)
        
        if status:
            query += " AND r.status = ?"
            params.append(status.value)
        
        if year:
            query += " AND strftime('%Y', r.date) = ?"
            params.append(str(year))
        
        # Apply sorting
        if sort_by == "date":
            query += f" ORDER BY r.date {sort_order.upper()}"
        elif sort_by == "asset":
            query += f" ORDER BY a.name {sort_order.upper()}"
        else:
            query += " ORDER BY r.date DESC"
        
        cursor.execute(query, params)
        repairs = []
        
        for row in cursor.fetchall():
            repairs.append(RepairResponse(
                id=row["id"],
                property_id=row["property_id"],
                date=datetime.fromisoformat(row["date"]).date(),
                description=row["description"],
                performed_by=row["performed_by"],
                notes=row["notes"],
                cost_cents=row["cost_cents"],
                status=row["status"],
                created_by_id=row["created_by_id"],
                created_at=datetime.fromisoformat(row["created_at"]),
                updated_at=datetime.fromisoformat(row["updated_at"])
            ))
        
        return repairs
    finally:
        conn.close()

@app.post("/api/repairs", response_model=RepairResponse)
async def create_repair(
    repair: RepairCreate,
    current_user: UserResponse = Depends(get_current_user)
):
    """Create a new repair"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Verify asset exists and user has permission
        query = "SELECT * FROM assets WHERE id = ?"
        params = [repair.property_id]
        
        if current_user.role != "ADMIN":
            query += " AND owner_id = ?"
            params.append(current_user.id)
        
        cursor.execute(query, params)
        asset_row = cursor.fetchone()
        
        if not asset_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")
        
        # Create repair
        cursor.execute(
            """INSERT INTO repairs 
               (property_id, date, description, performed_by, notes, cost_cents, status, created_by_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (repair.property_id, repair.date.isoformat(), repair.description,
             repair.performed_by, repair.notes, repair.cost_cents, repair.status.value, current_user.id)
        )
        conn.commit()
        
        repair_id = cursor.lastrowid
        cursor.execute("SELECT * FROM repairs WHERE id = ?", (repair_id,))
        repair_row = cursor.fetchone()
        
        return RepairResponse(
            id=repair_row["id"],
            property_id=repair_row["property_id"],
            date=datetime.fromisoformat(repair_row["date"]).date(),
            description=repair_row["description"],
            performed_by=repair_row["performed_by"],
            notes=repair_row["notes"],
            cost_cents=repair_row["cost_cents"],
            status=repair_row["status"],
            created_by_id=repair_row["created_by_id"],
            created_at=datetime.fromisoformat(repair_row["created_at"]),
            updated_at=datetime.fromisoformat(repair_row["updated_at"])
        )
    except Exception as e:
        conn.rollback()
        logger.error(f"Repair creation error: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Repair creation failed")
    finally:
        conn.close()

@app.get("/api/repairs/{repair_id}", response_model=RepairResponse)
async def get_repair(
    repair_id: int,
    current_user: UserResponse = Depends(get_current_user)
):
    """Get repair by ID"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        query = """
            SELECT r.* FROM repairs r
            JOIN assets a ON r.property_id = a.id
            WHERE r.id = ?
        """
        params = [repair_id]
        
        if current_user.role != "ADMIN":
            query += " AND a.owner_id = ?"
            params.append(current_user.id)
        
        cursor.execute(query, params)
        repair_row = cursor.fetchone()
        
        if not repair_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Repair not found")
        
        return RepairResponse(
            id=repair_row["id"],
            property_id=repair_row["property_id"],
            date=datetime.fromisoformat(repair_row["date"]).date(),
            description=repair_row["description"],
            performed_by=repair_row["performed_by"],
            notes=repair_row["notes"],
            cost_cents=repair_row["cost_cents"],
            status=repair_row["status"],
            created_by_id=repair_row["created_by_id"],
            created_at=datetime.fromisoformat(repair_row["created_at"]),
            updated_at=datetime.fromisoformat(repair_row["updated_at"])
        )
    finally:
        conn.close()

@app.put("/api/repairs/{repair_id}", response_model=RepairResponse)
async def update_repair(
    repair_id: int,
    repair_update: RepairUpdate,
    current_user: UserResponse = Depends(get_current_user)
):
    """Update repair"""
    logger.info(f"Updating repair {repair_id} with data: {repair_update}")
    logger.info(f"Repair update type: {type(repair_update)}")
    logger.info(f"Repair update dict: {repair_update.model_dump()}")
    
    # Convert date string to date object if provided
    if repair_update.date:
        from datetime import datetime
        try:
            repair_update.date = datetime.strptime(repair_update.date, '%Y-%m-%d').date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Expected YYYY-MM-DD")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if repair exists and user has permission
        query = """
            SELECT r.* FROM repairs r
            JOIN assets a ON r.property_id = a.id
            WHERE r.id = ?
        """
        params = [repair_id]
        
        if current_user.role != "ADMIN":
            query += " AND a.owner_id = ?"
            params.append(current_user.id)
        
        cursor.execute(query, params)
        repair_row = cursor.fetchone()
        
        if not repair_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Repair not found")
        
        # Update repair - update all provided fields
        update_fields = []
        update_params = []
        
        # Always update all fields that are provided
        if repair_update.property_id is not None:
            # Verify new asset exists and user has permission
            asset_query = "SELECT * FROM assets WHERE id = ?"
            asset_params = [repair_update.property_id]
            
            if current_user.role != "ADMIN":
                asset_query += " AND owner_id = ?"
                asset_params.append(current_user.id)
            
            cursor.execute(asset_query, asset_params)
            if not cursor.fetchone():
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")
            
            update_fields.append("property_id = ?")
            update_params.append(repair_update.property_id)
        
        if repair_update.date is not None:
            update_fields.append("date = ?")
            update_params.append(repair_update.date.isoformat())
        
        if repair_update.description is not None:
            update_fields.append("description = ?")
            update_params.append(repair_update.description)
        
        if repair_update.performed_by is not None:
            update_fields.append("performed_by = ?")
            update_params.append(repair_update.performed_by)
        
        # Update notes field (can be None for empty notes)
        update_fields.append("notes = ?")
        update_params.append(repair_update.notes)
        
        if repair_update.cost_cents is not None:
            update_fields.append("cost_cents = ?")
            update_params.append(repair_update.cost_cents)
        
        if repair_update.status is not None:
            update_fields.append("status = ?")
            update_params.append(repair_update.status.value)
        
        if update_fields:
            update_fields.append("updated_at = ?")
            update_params.append(datetime.utcnow().isoformat())
            update_params.append(repair_id)
            
            cursor.execute(
                f"UPDATE repairs SET {', '.join(update_fields)} WHERE id = ?",
                update_params
            )
            conn.commit()
        
        # Return updated repair
        cursor.execute("SELECT * FROM repairs WHERE id = ?", (repair_id,))
        updated_repair = cursor.fetchone()
        
        return RepairResponse(
            id=updated_repair["id"],
            property_id=updated_repair["property_id"],
            date=datetime.fromisoformat(updated_repair["date"]).date(),
            description=updated_repair["description"],
            performed_by=updated_repair["performed_by"],
            notes=updated_repair["notes"],
            cost_cents=updated_repair["cost_cents"],
            status=updated_repair["status"],
            created_by_id=updated_repair["created_by_id"],
            created_at=datetime.fromisoformat(updated_repair["created_at"]),
            updated_at=datetime.fromisoformat(updated_repair["updated_at"])
        )
    except Exception as e:
        conn.rollback()
        logger.error(f"Repair update error: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Repair update failed")
    finally:
        conn.close()

@app.delete("/api/repairs/{repair_id}")
async def delete_repair(
    repair_id: int,
    current_user: UserResponse = Depends(get_current_user)
):
    """Delete repair"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if repair exists and user has permission
        query = """
            SELECT r.* FROM repairs r
            JOIN assets a ON r.property_id = a.id
            WHERE r.id = ?
        """
        params = [repair_id]
        
        if current_user.role != "ADMIN":
            query += " AND a.owner_id = ?"
            params.append(current_user.id)
        
        cursor.execute(query, params)
        repair_row = cursor.fetchone()
        
        if not repair_row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Repair not found")
        
        # Delete repair
        cursor.execute("DELETE FROM repairs WHERE id = ?", (repair_id,))
        conn.commit()
        
        return {"message": "Repair deleted successfully"}
    except Exception as e:
        conn.rollback()
        logger.error(f"Repair deletion error: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Repair deletion failed")
    finally:
        conn.close()

# Export endpoints
@app.get("/api/export/repairs")
async def export_repairs(
    current_user: UserResponse = Depends(get_current_user),
    property_id: Optional[int] = None,
    status: Optional[RepairStatus] = None,
    year: Optional[int] = None
):
    """Export repairs to Excel"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Build query with same filtering as get_repairs
        query = """
            SELECT r.*, a.name as asset_name, a.type as asset_type
            FROM repairs r
            JOIN assets a ON r.property_id = a.id
            WHERE 1=1
        """
        params = []
        
        if current_user.role != "ADMIN":
            query += " AND a.owner_id = ?"
            params.append(current_user.id)
        
        if property_id:
            query += " AND r.property_id = ?"
            params.append(property_id)
        
        if status:
            query += " AND r.status = ?"
            params.append(status.value)
        
        if year:
            query += " AND strftime('%Y', r.date) = ?"
            params.append(str(year))
        
        query += " ORDER BY r.date DESC"
        
        cursor.execute(query, params)
        repairs = cursor.fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Repairs"
        
        # Headers
        headers = [
            "ID", "Asset Name", "Asset Type", "Date", "Description", 
            "Performed By", "Notes", "Cost (cents)", "Status", "Created At"
        ]
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data
        for row, repair in enumerate(repairs, 2):
            ws.cell(row=row, column=1, value=repair["id"])
            ws.cell(row=row, column=2, value=repair["asset_name"])
            ws.cell(row=row, column=3, value=repair["asset_type"])
            ws.cell(row=row, column=4, value=repair["date"])
            ws.cell(row=row, column=5, value=repair["description"])
            ws.cell(row=row, column=6, value=repair["performed_by"])
            ws.cell(row=row, column=7, value=repair["notes"])
            ws.cell(row=row, column=8, value=repair["cost_cents"])
            ws.cell(row=row, column=9, value=repair["status"])
            ws.cell(row=row, column=10, value=repair["created_at"])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create a temporary file for the response
        import tempfile
        import os
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(output.read())
            tmp_file_path = tmp_file.name
        
        return FileResponse(
            tmp_file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"repairs_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    finally:
        conn.close()

@app.get("/api/export/assets")
async def export_assets(
    current_user: UserResponse = Depends(get_current_user),
    owner_id: Optional[int] = None,
    type: Optional[str] = None
):
    """Export assets to Excel"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        query = """
            SELECT a.*, u.name as owner_name, u.email as owner_email
            FROM assets a
            LEFT JOIN users u ON a.owner_id = u.id
            WHERE 1=1
        """
        params = []
        
        if current_user.role != "ADMIN":
            query += " AND a.owner_id = ?"
            params.append(current_user.id)
        elif owner_id is not None:
            query += " AND a.owner_id = ?"
            params.append(owner_id)
        
        if type:
            query += " AND a.type = ?"
            params.append(type)
        
        query += " ORDER BY a.created_at DESC"
        
        cursor.execute(query, params)
        assets = cursor.fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Assets"
        
        # Headers
        headers = [
            "ID", "Name", "Type", "Owner Name", "Owner Email", "Created At", "Updated At"
        ]
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data
        for row, asset in enumerate(assets, 2):
            ws.cell(row=row, column=1, value=asset["id"])
            ws.cell(row=row, column=2, value=asset["name"])
            ws.cell(row=row, column=3, value=asset["type"])
            ws.cell(row=row, column=4, value=asset["owner_name"])
            ws.cell(row=row, column=5, value=asset["owner_email"])
            ws.cell(row=row, column=6, value=asset["created_at"])
            ws.cell(row=row, column=7, value=asset["updated_at"])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create a temporary file for the response
        import tempfile
        import os
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(output.read())
            tmp_file_path = tmp_file.name
        
        return FileResponse(
            tmp_file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"assets_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    finally:
        conn.close()

# Import endpoints
@app.post("/api/import/assets", response_model=ImportResult)
async def import_assets_csv(
    file: UploadFile = File(...),
    current_user: UserResponse = Depends(get_current_user)
):
    """Import assets from CSV file with semicolon separator"""
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="File must be a CSV file")
    
    result = ImportResult(total_rows=0, successful_imports=0, failed_imports=0, errors=[])
    
    try:
        # Read CSV content
        content = await file.read()
        csv_content = detect_encoding(content)
        
        # Parse CSV with semicolon separator
        csv_reader = csv.DictReader(io.StringIO(csv_content), delimiter=';')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 because of header
                result.total_rows += 1
                
                try:
                    # Validate row data
                    asset_data = AssetImportRow(**row)
                    
                    # Check if asset already exists for this user
                    cursor.execute(
                        "SELECT id FROM assets WHERE name = ? AND owner_id = ?",
                        (asset_data.name, current_user.id)
                    )
                    
                    if cursor.fetchone():
                        result.failed_imports += 1
                        result.errors.append(f"Row {row_num}: Asset '{asset_data.name}' already exists")
                        continue
                    
                    # Create asset
                    cursor.execute(
                        "INSERT INTO assets (name, type, owner_id) VALUES (?, ?, ?)",
                        (asset_data.name, asset_data.type, current_user.id)
                    )
                    result.successful_imports += 1
                    
                except Exception as e:
                    result.failed_imports += 1
                    result.errors.append(f"Row {row_num}: {str(e)}")
            
            conn.commit()
            
        except Exception as e:
            conn.rollback()
            raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
        finally:
            conn.close()
            
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing CSV file: {str(e)}")
    
    return result

@app.post("/api/import/repairs", response_model=ImportResult)
async def import_repairs_csv(
    file: UploadFile = File(...),
    current_user: UserResponse = Depends(get_current_user)
):
    """Import repairs from CSV file with semicolon separator"""
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="File must be a CSV file")
    
    result = ImportResult(total_rows=0, successful_imports=0, failed_imports=0, errors=[])
    
    try:
        # Read CSV content
        content = await file.read()
        csv_content = detect_encoding(content)
        
        # Parse CSV with semicolon separator
        csv_reader = csv.DictReader(io.StringIO(csv_content), delimiter=';')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Get all user's assets for lookup
            cursor.execute("SELECT id, name FROM assets WHERE owner_id = ?", (current_user.id,))
            user_assets = {row["name"]: row["id"] for row in cursor.fetchall()}
            
            for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 because of header
                result.total_rows += 1
                
                try:
                    # Validate row data
                    repair_data = RepairImportRow(**row)
                    
                    # Check if asset exists
                    if repair_data.asset_name not in user_assets:
                        result.failed_imports += 1
                        result.errors.append(f"Row {row_num}: Asset '{repair_data.asset_name}' not found")
                        continue
                    
                    asset_id = user_assets[repair_data.asset_name]
                    
                    # Create repair
                    cursor.execute(
                        """INSERT INTO repairs 
                           (property_id, date, description, performed_by, notes, cost_cents, status, created_by_id)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                        (asset_id, repair_data.date.isoformat(), repair_data.description,
                         repair_data.performed_by, repair_data.notes, repair_data.cost_cents, 
                         repair_data.status, current_user.id)
                    )
                    result.successful_imports += 1
                    
                except Exception as e:
                    result.failed_imports += 1
                    result.errors.append(f"Row {row_num}: {str(e)}")
            
            conn.commit()
            
        except Exception as e:
            conn.rollback()
            raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
        finally:
            conn.close()
            
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing CSV file: {str(e)}")
    
    return result

# Template download endpoints
@app.get("/api/templates/assets")
async def download_assets_template():
    """Download CSV template for assets import"""
    # Create CSV content
    csv_content = "name;type\n"
    csv_content += "House on Main Street;PROPERTY\n"
    csv_content += "Office Building;PROPERTY\n"
    csv_content += "Company Car;VEHICLE\n"
    
    # Create file-like object
    file_like = io.StringIO(csv_content)
    
    return StreamingResponse(
        io.BytesIO(csv_content.encode('utf-8')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=assets_template.csv"}
    )

@app.get("/api/templates/repairs")
async def download_repairs_template():
    """Download CSV template for repairs import"""
    # Create CSV content
    csv_content = "asset_name;date;description;performed_by;notes;cost_cents;status\n"
    csv_content += "House on Main Street;2024-01-15;Roof repair;John Smith;Fixed leak in roof;643.36;COMPLETED\n"
    csv_content += "Office Building;2024-02-10;Heating maintenance;Mike Johnson;Annual service;150.50;COMPLETED\n"
    csv_content += "Company Car;2024-03-05;Oil change;Auto Service;Regular maintenance;75.00;COMPLETED\n"
    
    # Create file-like object
    file_like = io.StringIO(csv_content)
    
    return StreamingResponse(
        io.BytesIO(csv_content.encode('utf-8')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=repairs_template.csv"}
    )

# Profile endpoints
@app.get("/api/profile", response_model=UserResponse)
async def get_profile(current_user: UserResponse = Depends(get_current_user)):
    """Get current user profile"""
    return current_user

@app.put("/api/profile", response_model=UserResponse)
async def update_profile(
    profile_update: UserProfileUpdate,
    current_user: UserResponse = Depends(get_current_user)
):
    """Update user profile"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if email is already taken by another user
        if profile_update.email and profile_update.email != current_user.email:
            cursor.execute("SELECT id FROM users WHERE email = ? AND id != ?", 
                         (profile_update.email, current_user.id))
            if cursor.fetchone():
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Email already taken"
                )
        
        # Update user
        update_fields = []
        update_params = []
        
        if profile_update.name is not None:
            update_fields.append("name = ?")
            update_params.append(profile_update.name)
        
        if profile_update.email is not None:
            update_fields.append("email = ?")
            update_params.append(profile_update.email)
        
        if update_fields:
            update_fields.append("updated_at = ?")
            update_params.append(datetime.utcnow().isoformat())
            update_params.append(current_user.id)
            
            cursor.execute(
                f"UPDATE users SET {', '.join(update_fields)} WHERE id = ?",
                update_params
            )
            conn.commit()
        
        # Return updated user
        cursor.execute("SELECT * FROM users WHERE id = ?", (current_user.id,))
        updated_user = cursor.fetchone()
        
        return UserResponse(
            id=updated_user["id"],
            email=updated_user["email"],
            name=updated_user["name"],
            role=updated_user["role"],
            created_at=datetime.fromisoformat(updated_user["created_at"]),
            updated_at=datetime.fromisoformat(updated_user["updated_at"])
        )
    except Exception as e:
        conn.rollback()
        logger.error(f"Profile update error: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Profile update failed")
    finally:
        conn.close()

@app.post("/api/profile/change-password")
async def change_password(
    password_data: PasswordChange,
    current_user: UserResponse = Depends(get_current_user)
):
    """Change user password"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Verify current password
        cursor.execute("SELECT password_hash FROM users WHERE id = ?", (current_user.id,))
        password_row = cursor.fetchone()
        
        if not password_row or not verify_password(password_data.current_password, password_row["password_hash"]):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Current password is incorrect"
            )
        
        # Update password
        new_password_hash = get_password_hash(password_data.new_password)
        cursor.execute(
            "UPDATE users SET password_hash = ?, updated_at = ? WHERE id = ?",
            (new_password_hash, datetime.utcnow().isoformat(), current_user.id)
        )
        conn.commit()
        
        return {"message": "Password changed successfully"}
    except Exception as e:
        conn.rollback()
        logger.error(f"Password change error: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Password change failed")
    finally:
        conn.close()

@app.get("/api/profile/settings", response_model=UserSettingsResponse)
async def get_user_settings(current_user: UserResponse = Depends(get_current_user)):
    """Get user settings"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT * FROM user_settings WHERE user_id = ?", (current_user.id,))
        settings_row = cursor.fetchone()
        
        if not settings_row:
            # Create default settings if they don't exist
            cursor.execute(
                "INSERT INTO user_settings (user_id, currency, language, date_format, theme) VALUES (?, ?, ?, ?, ?)",
                (current_user.id, "USD", "en", "DD.MM.YYYY", "dark")
            )
            conn.commit()
            
            cursor.execute("SELECT * FROM user_settings WHERE user_id = ?", (current_user.id,))
            settings_row = cursor.fetchone()
        
        return UserSettingsResponse(
            id=settings_row["id"],
            user_id=settings_row["user_id"],
            currency=settings_row["currency"],
            language=settings_row["language"],
            date_format=settings_row["date_format"],
            theme=settings_row["theme"],
            created_at=datetime.fromisoformat(settings_row["created_at"]),
            updated_at=datetime.fromisoformat(settings_row["updated_at"])
        )
    finally:
        conn.close()

@app.put("/api/profile/settings", response_model=UserSettingsResponse)
async def update_user_settings(
    settings: UserSettings,
    current_user: UserResponse = Depends(get_current_user)
):
    """Update user settings"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if settings exist
        cursor.execute("SELECT id FROM user_settings WHERE user_id = ?", (current_user.id,))
        settings_row = cursor.fetchone()
        
        if settings_row:
            # Update existing settings
            cursor.execute(
                """UPDATE user_settings 
                   SET currency = ?, language = ?, date_format = ?, theme = ?, updated_at = ?
                   WHERE user_id = ?""",
                (settings.currency, settings.language, settings.date_format, 
                 settings.theme, datetime.utcnow().isoformat(), current_user.id)
            )
        else:
            # Create new settings
            cursor.execute(
                """INSERT INTO user_settings (user_id, currency, language, date_format, theme)
                   VALUES (?, ?, ?, ?, ?)""",
                (current_user.id, settings.currency, settings.language, 
                 settings.date_format, settings.theme)
            )
        
        conn.commit()
        
        # Return updated settings
        cursor.execute("SELECT * FROM user_settings WHERE user_id = ?", (current_user.id,))
        updated_settings = cursor.fetchone()
        
        return UserSettingsResponse(
            id=updated_settings["id"],
            user_id=updated_settings["user_id"],
            currency=updated_settings["currency"],
            language=updated_settings["language"],
            date_format=updated_settings["date_format"],
            theme=updated_settings["theme"],
            created_at=datetime.fromisoformat(updated_settings["created_at"]),
            updated_at=datetime.fromisoformat(updated_settings["updated_at"])
        )
    except Exception as e:
        conn.rollback()
        logger.error(f"Settings update error: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Settings update failed")
    finally:
        conn.close()

# Admin endpoints
@app.get("/api/admin/users", response_model=List[UserResponse])
async def get_users(current_user: UserResponse = Depends(get_current_admin_user)):
    """Get all users (admin only)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT * FROM users ORDER BY created_at DESC")
        users = []
        
        for row in cursor.fetchall():
            users.append(UserResponse(
                id=row["id"],
                email=row["email"],
                name=row["name"],
                role=row["role"],
                created_at=datetime.fromisoformat(row["created_at"]),
                updated_at=datetime.fromisoformat(row["updated_at"])
            ))
        
        return users
    finally:
        conn.close()

# Web interface routes
@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    """Dashboard page"""
    return templates.TemplateResponse("dashboard.html", {"request": request})

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """Login page"""
    return templates.TemplateResponse("login.html", {"request": request})

@app.get("/register", response_class=HTMLResponse)
async def register_page(request: Request):
    """Register page"""
    return templates.TemplateResponse("register.html", {"request": request})

@app.get("/assets", response_class=HTMLResponse)
async def assets_page(request: Request):
    """Assets page"""
    return templates.TemplateResponse("assets/list.html", {"request": request})

@app.get("/assets/new", response_class=HTMLResponse)
async def new_asset_page(request: Request):
    """New asset page"""
    return templates.TemplateResponse("assets/new.html", {"request": request})

@app.get("/assets/{asset_id}", response_class=HTMLResponse)
async def asset_detail_page(request: Request, asset_id: int):
    """Asset detail page"""
    return templates.TemplateResponse("assets/detail.html", {"request": request, "asset_id": asset_id})

@app.get("/assets/{asset_id}/edit", response_class=HTMLResponse)
async def edit_asset_page(request: Request, asset_id: int):
    """Edit asset page"""
    return templates.TemplateResponse("assets/edit.html", {"request": request, "asset_id": asset_id})

@app.get("/repairs", response_class=HTMLResponse)
async def repairs_page(request: Request):
    """Repairs page"""
    return templates.TemplateResponse("repairs/list.html", {"request": request})

@app.get("/repairs/new", response_class=HTMLResponse)
async def new_repair_page(request: Request):
    """New repair page"""
    return templates.TemplateResponse("repairs/new.html", {"request": request})

@app.get("/repairs/{repair_id}/edit", response_class=HTMLResponse)
async def edit_repair_page(request: Request, repair_id: int):
    """Edit repair page"""
    return templates.TemplateResponse("repairs/edit.html", {"request": request, "repair_id": repair_id})

@app.get("/profile", response_class=HTMLResponse)
async def profile_page(request: Request):
    """Profile page"""
    return templates.TemplateResponse("profile.html", {"request": request})

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

